#!/usr/bin/env python3
import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo


BJT = ZoneInfo("Asia/Shanghai")


def parse_iso(value):
    if not value:
        return None
    text = str(value).replace("Z", "+00:00")
    return datetime.fromisoformat(text).astimezone(timezone.utc)


def iso_z(dt):
    return dt.astimezone(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")


def fmt_bjt(dt):
    return dt.astimezone(BJT).strftime("%Y-%m-%d %H:%M:%S")


def tsv_escape(value):
    return "" if value is None else str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ")


def load_input(path):
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(data, list):
        return {"quotes": data}
    return data


def normalize(data):
    original_dt = parse_iso(data.get("original_published_at_utc"))
    if not original_dt:
        raise SystemExit("Input JSON must include original_published_at_utc")

    extracted_dt = parse_iso(data.get("extracted_at_utc")) or datetime.now(timezone.utc)
    rows = []
    for index, quote in enumerate(data.get("quotes", []), start=1):
        quote_dt = parse_iso(quote.get("quote_published_at_utc") or quote.get("datetime"))
        if not quote_dt:
            continue
        views = int(quote.get("views") or 0)
        likes = int(quote.get("likes") or 0)
        hours_after = (quote_dt - original_dt).total_seconds() / 3600
        ratio = likes / views if views else 0
        rows.append({
            "index": index,
            "name": quote.get("name") or "",
            "handle": quote.get("handle") or "",
            "quote_url": quote.get("quote_url") or quote.get("url") or "",
            "quote_published_at_utc": iso_z(quote_dt),
            "quote_published_at_local": fmt_bjt(quote_dt),
            "hours_after_original": round(hours_after, 2),
            "within_24h": 0 <= hours_after <= 24,
            "within_48h": 0 <= hours_after <= 48,
            "views": views,
            "likes": likes,
            "engage_likes_per_view": round(ratio, 4),
            "engage_likes_per_view_pct": f"{ratio * 100:.2f}%",
            "replies": int(quote.get("replies") or 0),
            "reposts": int(quote.get("reposts") or 0),
            "stat_label": quote.get("stat_label") or quote.get("statLabel") or "",
            "text_preview": quote.get("text_preview") or "",
        })

    rows.sort(key=lambda row: row["quote_published_at_utc"], reverse=True)
    for index, row in enumerate(rows, start=1):
        row["index"] = index

    age_hours = (extracted_dt - original_dt).total_seconds() / 3600
    summary = {
        "original_url": data.get("original_url") or "",
        "original_published_at_utc": iso_z(original_dt),
        "original_published_at_local": fmt_bjt(original_dt),
        "extracted_at_utc": iso_z(extracted_dt),
        "extracted_at_local": fmt_bjt(extracted_dt),
        "age_hours_at_extraction": round(age_hours, 2),
        "quote_count_total": len(rows),
        "quote_count_within_24h": sum(1 for row in rows if row["within_24h"]),
        "quote_count_within_48h": sum(1 for row in rows if row["within_48h"]),
        "window_note": (
            "24h and 48h windows are not final yet."
            if age_hours < 24
            else "24h is final; 48h is not final yet."
            if age_hours < 48
            else "24h and 48h windows are final."
        ),
    }
    return summary, rows


def write_tsv(path, rows):
    headers = [
        "#", "name", "handle", "quote_url", "quote_published_at_local", "hours_after_original",
        "within_24h", "within_48h", "views", "likes", "Engage(likes)/View",
        "Engage(likes)/View %", "replies", "reposts", "stat_label", "text_preview",
    ]
    lines = ["\t".join(headers)]
    for row in rows:
        values = [
            row["index"], row["name"], row["handle"], row["quote_url"], row["quote_published_at_local"],
            row["hours_after_original"], "Y" if row["within_24h"] else "N",
            "Y" if row["within_48h"] else "N", row["views"], row["likes"],
            row["engage_likes_per_view"], row["engage_likes_per_view_pct"], row["replies"],
            row["reposts"], row["stat_label"], row["text_preview"],
        ]
        lines.append("\t".join(tsv_escape(value) for value in values))
    path.write_text("\n".join(lines), encoding="utf-8")


def write_xlsx(path, summary, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    for key, value in summary.items():
        ws.append([key, value])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 92

    headers = [
        "#", "name", "handle", "quote_url", "quote_published_at_local", "hours_after_original",
        "within_24h", "within_48h", "views", "likes", "Engage(likes)/View",
        "Engage(likes)/View %", "replies", "reposts", "stat_label", "text_preview",
    ]
    widths = [5, 22, 18, 52, 22, 18, 12, 12, 12, 10, 18, 18, 10, 10, 48, 80]
    fill = PatternFill("solid", fgColor="1F4E78")

    def add_sheet(title, sheet_rows):
        sheet = wb.create_sheet(title)
        sheet.append(headers)
        for row in sheet_rows:
            sheet.append([
                row["index"], row["name"], row["handle"], row["quote_url"],
                row["quote_published_at_local"], row["hours_after_original"],
                "Y" if row["within_24h"] else "N", "Y" if row["within_48h"] else "N",
                row["views"], row["likes"], row["engage_likes_per_view"],
                row["engage_likes_per_view_pct"], row["replies"], row["reposts"],
                row["stat_label"], row["text_preview"],
            ])
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(wrap_text=True)
        sheet.freeze_panes = "A2"
        for i, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = width
        for body_row in sheet.iter_rows(min_row=2):
            for cell in body_row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    add_sheet("Quote Tracking", rows)
    add_sheet("By Views", sorted(rows, key=lambda row: row["views"], reverse=True))
    wb.save(path)
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--basename", default="x_quote_tracking")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    data = load_input(args.input)
    summary, rows = normalize(data)
    base = output_dir / f"{args.basename}_quote_tracking"

    json_path = base.with_suffix(".json")
    tsv_path = base.with_suffix(".tsv")
    xlsx_path = base.with_suffix(".xlsx")

    json_path.write_text(json.dumps({"summary": summary, "rows": rows}, indent=2, ensure_ascii=False), encoding="utf-8")
    write_tsv(tsv_path, rows)
    wrote_xlsx = write_xlsx(xlsx_path, summary, rows)

    result = {"json": str(json_path), "tsv": str(tsv_path), "xlsx": str(xlsx_path) if wrote_xlsx else None, "summary": summary}
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
